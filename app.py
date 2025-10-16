import streamlit as st
import pandas as pd
import json
from datetime import datetime
import io
import sys

# Import avec gestion d'erreur pour openpyxl
try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils.dataframe import dataframe_to_rows
    OPENPYXL_AVAILABLE = True
except ImportError as e:
    st.error(f"⚠️ Erreur d'import openpyxl: {e}")
    st.info("📦 Installation de openpyxl en cours...")
    try:
        import subprocess
        subprocess.check_call([sys.executable, "-m", "pip", "install", "openpyxl==3.1.5", "--no-cache-dir"])
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils.dataframe import dataframe_to_rows
        OPENPYXL_AVAILABLE = True
        st.success("✅ openpyxl installé avec succès!")
    except Exception as install_error:
        st.error(f"❌ Impossible d'installer openpyxl: {install_error}")
        OPENPYXL_AVAILABLE = False

# Configuration de la page
st.set_page_config(
    page_title="Audit BIOCOOP - Fournisseurs Locaux",
    page_icon="🏪",
    layout="wide",
    initial_sidebar_state="expanded"
)

# Définition de la checklist d'audit basée sur le cahier des charges
CHECKLIST_AUDIT = {
    "1. SÉCURITÉ DES ALIMENTS": {
        "criticite": "CRITIQUE",
        "coefficient": 2.0,
        "items": [
            {
                "id": "SEC-001",
                "question": "Pertinence et vérification des points critiques (étude HACCP) si applicable",
                "details": "Vérifier l'existence et la mise à jour du plan HACCP, identification des CCP"
            },
            {
                "id": "SEC-002",
                "question": "Maîtrise du risque allergènes",
                "details": "Procédures de gestion des allergènes, étiquetage, formation du personnel"
            },
            {
                "id": "SEC-003",
                "question": "Maîtrise du risque corps étrangers",
                "details": "Procédures de prévention, détection (tamis, aimants, détecteur de métaux)"
            },
            {
                "id": "SEC-004",
                "question": "Maîtrise du risque chimique",
                "details": "Stockage et utilisation des produits chimiques, traçabilité"
            },
            {
                "id": "SEC-005",
                "question": "Définition d'un plan d'analyses et d'un plan de contrôles",
                "details": "Plan d'analyses microbiologiques, physico-chimiques, fréquence, laboratoire"
            },
            {
                "id": "SEC-006",
                "question": "Système de blocage/libération produits si concerné",
                "details": "Procédure de quarantaine et de libération des produits"
            },
            {
                "id": "SEC-007",
                "question": "Gestion des retraits rappels produits (traçabilité des lots livrés aux magasins)",
                "details": "Procédure de retrait/rappel, traçabilité amont/aval, tests de traçabilité"
            }
        ]
    },
    "2. HYGIÈNE DU PERSONNEL ET DES LOCAUX": {
        "criticite": "CRITIQUE",
        "coefficient": 2.0,
        "items": [
            {
                "id": "HYG-001",
                "question": "Définition/affichage des règles d'hygiène",
                "details": "Règles affichées, accessibles et compréhensibles par le personnel"
            },
            {
                "id": "HYG-002",
                "question": "Respect application des règles d'hygiène par le personnel",
                "details": "Port de la tenue, lavage des mains, comportement en production"
            },
            {
                "id": "HYG-003",
                "question": "Conformité infrastructure (locaux de production, sanitaires, vestiaires, équipements)",
                "details": "État des locaux, séparation des zones, équipements pour l'hygiène"
            },
            {
                "id": "HYG-004",
                "question": "Maîtrise des opérations de nettoyage",
                "details": "Plan de nettoyage, produits utilisés, fréquence, enregistrements"
            },
            {
                "id": "HYG-005",
                "question": "Maîtrise des températures (process et installation)",
                "details": "Contrôle des températures, enregistrements, actions correctives"
            },
            {
                "id": "HYG-006",
                "question": "Surveillance des nuisibles",
                "details": "Plan de lutte contre les nuisibles, prestataire, fréquence, résultats"
            }
        ]
    },
    "3. RÉGLEMENTATION": {
        "criticite": "MAJEUR",
        "coefficient": 1.5,
        "items": [
            {
                "id": "REG-001",
                "question": "Certificat bio couvrant l'ensemble des produits en cours de validité",
                "details": "Vérifier la validité du certificat bio et la couverture de tous les produits"
            },
            {
                "id": "REG-002",
                "question": "Maîtrise des contrôles quantitatifs (poids/volume)",
                "details": "Procédure de contrôle des poids/volumes, balance étalonnée"
            },
            {
                "id": "REG-003",
                "question": "Conformité globale des étiquettes produits + marquage lot/DLC",
                "details": "Étiquetage réglementaire, liste d'ingrédients, allergènes, lot, DLC/DLUO"
            },
            {
                "id": "REG-004",
                "question": "Conformité système de traçabilité produits",
                "details": "Traçabilité amont/aval, test de traçabilité réalisé"
            }
        ]
    },
    "4. EXIGENCES BIOCOOP - PRODUITS AUTORISÉS": {
        "criticite": "MAJEUR",
        "coefficient": 1.5,
        "items": [
            {
                "id": "BIO-001",
                "question": "Produits certifiés bio portant le label européen OU en conversion OU SPG autorisé",
                "details": "Vérifier Nature et Progrès ou Simples, attestation en cours de validité"
            },
            {
                "id": "BIO-002",
                "question": "Produits de la mer : ingrédient principal non certifiable + ingrédients bio",
                "details": "Si applicable, vérifier conformité avec annexe produits de la pêche"
            },
            {
                "id": "BIO-003",
                "question": "Absence de dioxyde de silicium (E551) et dioxyde de titane (E171)",
                "details": "Vérifier les recettes et étiquettes"
            }
        ]
    },
    "5. EXIGENCES BIOCOOP - INGRÉDIENTS SPÉCIFIQUES": {
        "criticite": "MAJEUR",
        "coefficient": 1.5,
        "items": [
            {
                "id": "ING-001",
                "question": "Sel : de mer, récolté manuellement, origine France",
                "details": "Pour le sel vendu en l'état (aromatisé ou non)"
            },
            {
                "id": "ING-002",
                "question": "Arômes : certifiés biologiques",
                "details": "Si goût annoncé dans le nom, tous les ingrédients aromatisants doivent être bio"
            },
            {
                "id": "ING-003",
                "question": "Absence de labels interdits sur les étiquettes",
                "details": "HVE, Zéro résidu, Produit responsable, Bleu Blanc Cœur, Bee Friendly, etc."
            }
        ]
    },
    "6. EXIGENCES BIOCOOP - COMMERCE ÉQUITABLE": {
        "criticite": "MAJEUR",
        "coefficient": 1.5,
        "items": [
            {
                "id": "CEQ-001",
                "question": "Produits bruts obligatoirement commerce équitable : Café, Sucre de canne/coco, Chocolat",
                "details": "Vérifier label présent sur l'étiquette (liste des labels autorisés disponible)"
            },
            {
                "id": "CEQ-002",
                "question": "Thé (hors Japon/Corée du Sud), Beurre de cacao certifiés commerce équitable",
                "details": "Label présent sur l'étiquette"
            },
            {
                "id": "CEQ-003",
                "question": "Fruits secs mono-ingrédient : cajou, macadamia, coco, Brésil, ananas, papaye, banane, mangue",
                "details": "Commerce équitable obligatoire (vrac et conditionné)"
            },
            {
                "id": "CEQ-004",
                "question": "Gingembre confit (Hors UE), Beurre de karité, Huile d'argan, Riz hors UE",
                "details": "Commerce équitable obligatoire"
            },
            {
                "id": "CEQ-005",
                "question": "Pâtes à tartiner : sucre de canne commerce équitable avec mention",
                "details": "Mention '*issu du commerce équitable' en fin de liste d'ingrédients"
            }
        ]
    },
    "7. EXIGENCES BIOCOOP - TRANSPORT ET PRODUCTION": {
        "criticite": "MAJEUR",
        "coefficient": 1.5,
        "items": [
            {
                "id": "TRA-001",
                "question": "Interdiction transport aérien pour produits finis",
                "details": "Vérifier les modes de transport utilisés"
            },
            {
                "id": "TRA-002",
                "question": "Pas de serres chauffées (hors production de plants)",
                "details": "Pour fruits et légumes uniquement"
            },
            {
                "id": "TRA-003",
                "question": "Interdiction déverdissage des agrumes vendus en l'état",
                "details": "Vérifier les pratiques"
            },
            {
                "id": "TRA-004",
                "question": "Tomates 'anciennes' : variétés de population non hybride",
                "details": "Si applicable, vérifier les variétés"
            },
            {
                "id": "TRA-005",
                "question": "Produits de la pêche : conformité avec annexe zones autorisées",
                "details": "Cf. Liste des produits de la pêche et zones BIOCOOP"
            }
        ]
    },
    "8. MAÎTRISE ORIGINES MATIÈRES PREMIÈRES": {
        "criticite": "STANDARD",
        "coefficient": 1.0,
        "items": [
            {
                "id": "ORI-001",
                "question": "Connaissance des origines des ingrédients",
                "details": "Cohérence avec le tableau des origines BIOCOOP (par sondage)"
            },
            {
                "id": "ORI-002",
                "question": "Respect des couples produits/origines interdits",
                "details": "Ex: Ail de Chine, Sucre de canne d'Inde, Cacao RDC, Muscade Indonésie"
            },
            {
                "id": "ORI-003",
                "question": "Respect des zones obligatoires pour produits spécifiques",
                "details": "Ex: Agrumes bassin méditerranéen, Produits animaux UE, etc."
            }
        ]
    },
    "9. MAÎTRISE OGM": {
        "criticite": "STANDARD",
        "coefficient": 1.0,
        "items": [
            {
                "id": "OGM-001",
                "question": "Maîtrise de l'absence d'OGM dans les produits",
                "details": "Seuil analytique: 0,01% produits bruts, 0,1% ingrédients élaborés"
            },
            {
                "id": "OGM-002",
                "question": "Contrôle des ingrédients à risque OGM",
                "details": "Ananas rose, betterave, maïs, colza, papaye, sirop/riz basmati, soja, sucre de canne"
            }
        ]
    }
}

# Options de notation
NOTATION_OPTIONS = {
    "A": {"label": "A - Conforme", "points": 20, "color": "#28a745"},
    "B": {"label": "B - Non-conformité mineure", "points": 10, "color": "#ffc107"},
    "C": {"label": "C - Non-conformité majeure", "points": 0, "color": "#dc3545"},
    "N/A": {"label": "N/A - Non applicable", "points": None, "color": "#6c757d"}
}

def initialize_session_state():
    """Initialise l'état de la session"""
    if 'audit_data' not in st.session_state:
        st.session_state.audit_data = {}
    if 'fournisseur_info' not in st.session_state:
        st.session_state.fournisseur_info = {}
    if 'current_step' not in st.session_state:
        st.session_state.current_step = 1

def calculer_score_global(audit_data):
    """Calcule le score global de l'audit"""
    total_points = 0
    total_possible = 0
    details_par_categorie = {}
    
    for categorie, data in CHECKLIST_AUDIT.items():
        coefficient = data["coefficient"]
        points_categorie = 0
        points_possibles_categorie = 0
        items_evalues = 0
        
        for item in data["items"]:
            item_id = item["id"]
            if item_id in audit_data and audit_data[item_id]["notation"] != "N/A":
                note = audit_data[item_id]["notation"]
                points = NOTATION_OPTIONS[note]["points"]
                points_categorie += points * coefficient
                points_possibles_categorie += 20 * coefficient
                items_evalues += 1
        
        if items_evalues > 0:
            score_categorie = (points_categorie / points_possibles_categorie) * 100
            details_par_categorie[categorie] = {
                "score": score_categorie,
                "points": points_categorie,
                "points_possibles": points_possibles_categorie,
                "items_evalues": items_evalues,
                "criticite": data["criticite"]
            }
            total_points += points_categorie
            total_possible += points_possibles_categorie
    
    score_global = (total_points / total_possible * 100) if total_possible > 0 else 0
    
    return score_global, details_par_categorie

def get_niveau_conformite(score):
    """Détermine le niveau de conformité selon le score"""
    if score >= 90:
        return "EXCELLENT", "#28a745"
    elif score >= 75:
        return "SATISFAISANT", "#5cb85c"
    elif score >= 60:
        return "ACCEPTABLE", "#ffc107"
    elif score >= 40:
        return "INSUFFISANT", "#fd7e14"
    else:
        return "NON CONFORME", "#dc3545"

def generer_rapport_excel(fournisseur_info, audit_data):
    """Génère un rapport d'audit complet en Excel"""
    if not OPENPYXL_AVAILABLE:
        st.error("❌ Impossible de générer le rapport Excel : openpyxl n'est pas disponible")
        return None
    
    try:
        wb = Workbook()
        
        # Styles
        header_fill = PatternFill(start_color="2C3E50", end_color="2C3E50", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=12)
        categorie_fill = PatternFill(start_color="3498DB", end_color="3498DB", fill_type="solid")
        categorie_font = Font(bold=True, color="FFFFFF", size=11)
        
        conforme_fill = PatternFill(start_color="D5F4E6", end_color="D5F4E6", fill_type="solid")
        mineur_fill = PatternFill(start_color="FFF3CD", end_color="FFF3CD", fill_type="solid")
        majeur_fill = PatternFill(start_color="F8D7DA", end_color="F8D7DA", fill_type="solid")
        
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # FEUILLE 1: Informations générales
        ws1 = wb.active
        ws1.title = "Informations Fournisseur"
        
        ws1['A1'] = "RAPPORT D'AUDIT FOURNISSEUR BIOCOOP"
        ws1['A1'].font = Font(bold=True, size=16)
        ws1.merge_cells('A1:D1')
        
        row = 3
        for key, value in fournisseur_info.items():
            ws1[f'A{row}'] = key
            ws1[f'A{row}'].font = Font(bold=True)
            ws1[f'B{row}'] = value
            row += 1
        
        # FEUILLE 2: Résultats d'audit détaillés
        ws2 = wb.create_sheet("Résultats Audit")
        
        # En-têtes
        headers = ["ID", "Catégorie", "Question", "Notation", "Commentaire", "Criticité"]
        for col, header in enumerate(headers, 1):
            cell = ws2.cell(row=1, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = border
        
        row = 2
        for categorie, data in CHECKLIST_AUDIT.items():
            for item in data["items"]:
                item_id = item["id"]
                if item_id in audit_data:
                    notation = audit_data[item_id]["notation"]
                    commentaire = audit_data[item_id].get("commentaire", "")
                    
                    ws2.cell(row=row, column=1, value=item_id)
                    ws2.cell(row=row, column=2, value=categorie.split(".")[1].strip())
                    ws2.cell(row=row, column=3, value=item["question"])
                    ws2.cell(row=row, column=4, value=notation)
                    ws2.cell(row=row, column=5, value=commentaire)
                    ws2.cell(row=row, column=6, value=data["criticite"])
                    
                    # Coloration selon notation
                    if notation == "A":
                        fill = conforme_fill
                    elif notation == "B":
                        fill = mineur_fill
                    elif notation == "C":
                        fill = majeur_fill
                    else:
                        fill = None
                    
                    if fill:
                        for col in range(1, 7):
                            ws2.cell(row=row, column=col).fill = fill
                    
                    for col in range(1, 7):
                        ws2.cell(row=row, column=col).border = border
                    
                    row += 1
        
        # Ajuster les largeurs de colonnes
        ws2.column_dimensions['A'].width = 12
        ws2.column_dimensions['B'].width = 30
        ws2.column_dimensions['C'].width = 50
        ws2.column_dimensions['D'].width = 12
        ws2.column_dimensions['E'].width = 40
        ws2.column_dimensions['F'].width = 15
        
        # FEUILLE 3: Plan d'action
        ws3 = wb.create_sheet("Plan d'Action")
        
        # En-têtes plan d'action
        action_headers = ["ID", "Point d'audit", "Non-conformité", "Action corrective", 
                         "Responsable", "Délai", "Statut", "Date de clôture", "Commentaires"]
        for col, header in enumerate(action_headers, 1):
            cell = ws3.cell(row=1, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.border = border
        
        row = 2
        for categorie, data in CHECKLIST_AUDIT.items():
            for item in data["items"]:
                item_id = item["id"]
                if item_id in audit_data and audit_data[item_id]["notation"] in ["B", "C"]:
                    ws3.cell(row=row, column=1, value=item_id)
                    ws3.cell(row=row, column=2, value=item["question"])
                    ws3.cell(row=row, column=3, value=audit_data[item_id].get("commentaire", ""))
                    ws3.cell(row=row, column=4, value="[À définir]")
                    ws3.cell(row=row, column=5, value="[Responsable]")
                    ws3.cell(row=row, column=6, value="[Date limite]")
                    ws3.cell(row=row, column=7, value="En cours")
                    ws3.cell(row=row, column=8, value="")
                    ws3.cell(row=row, column=9, value="")
                    
                    # Coloration selon criticité
                    fill = majeur_fill if audit_data[item_id]["notation"] == "C" else mineur_fill
                    for col in range(1, 10):
                        ws3.cell(row=row, column=col).fill = fill
                        ws3.cell(row=row, column=col).border = border
                    
                    row += 1
        
        # Ajuster les largeurs
        for col, width in zip(range(1, 10), [12, 40, 35, 35, 20, 15, 12, 15, 30]):
            ws3.column_dimensions[chr(64 + col)].width = width
        
        # FEUILLE 4: Synthèse et scores
        ws4 = wb.create_sheet("Synthèse")
        
        score_global, details = calculer_score_global(audit_data)
        niveau, couleur = get_niveau_conformite(score_global)
        
        ws4['A1'] = "SYNTHÈSE DE L'AUDIT"
        ws4['A1'].font = Font(bold=True, size=16)
        ws4.merge_cells('A1:D1')
        
        ws4['A3'] = "Score Global"
        ws4['B3'] = f"{score_global:.1f}%"
        ws4['A4'] = "Niveau de Conformité"
        ws4['B4'] = niveau
        
        ws4['A3'].font = Font(bold=True)
        ws4['A4'].font = Font(bold=True)
        ws4['B3'].font = Font(bold=True, size=14)
        ws4['B4'].font = Font(bold=True, size=14)
        
        # Scores par catégorie
        ws4['A6'] = "Scores par catégorie"
        ws4['A6'].font = Font(bold=True, size=12)
        
        row = 7
        ws4.cell(row=row, column=1, value="Catégorie")
        ws4.cell(row=row, column=2, value="Score (%)")
        ws4.cell(row=row, column=3, value="Criticité")
        ws4.cell(row=row, column=4, value="Items évalués")
        
        for col in range(1, 5):
            ws4.cell(row=row, column=col).fill = header_fill
            ws4.cell(row=row, column=col).font = header_font
        
        row += 1
        for categorie, info in details.items():
            ws4.cell(row=row, column=1, value=categorie.split(".")[1].strip())
            ws4.cell(row=row, column=2, value=f"{info['score']:.1f}%")
            ws4.cell(row=row, column=3, value=info['criticite'])
            ws4.cell(row=row, column=4, value=info['items_evalues'])
            row += 1
        
        ws4.column_dimensions['A'].width = 40
        ws4.column_dimensions['B'].width = 15
        ws4.column_dimensions['C'].width = 15
        ws4.column_dimensions['D'].width = 15
        
        # Sauvegarder dans un buffer
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        
        return buffer
    
    except Exception as e:
        st.error(f"❌ Erreur lors de la génération du rapport Excel : {e}")
        return None

# Interface principale
def main():
    initialize_session_state()
    
    # Sidebar
    with st.sidebar:
        st.markdown("### 🏪 BIOCOOP")
        st.markdown("*Audit Fournisseurs Locaux*")
        st.divider()
        
        st.title("Navigation")
        
        step = st.radio(
            "Étapes de l'audit",
            ["1️⃣ Informations Fournisseur", "2️⃣ Checklist d'Audit", "3️⃣ Rapport Final"],
            index=st.session_state.current_step - 1
        )
        
        if "1️⃣" in step:
            st.session_state.current_step = 1
        elif "2️⃣" in step:
            st.session_state.current_step = 2
        else:
            st.session_state.current_step = 3
        
        st.divider()
        st.markdown("### 📊 Système de notation")
        st.markdown("**A** (20 pts) = Conforme ✅")
        st.markdown("**B** (10 pts) = NC mineure ⚠️")
        st.markdown("**C** (0 pts) = NC majeure ❌")
        st.markdown("**N/A** = Non applicable ⊘")
        
        st.divider()
        st.caption("Version 1.0 - Octobre 2025")
    
    # Contenu principal
    if st.session_state.current_step == 1:
        afficher_etape_informations()
    elif st.session_state.current_step == 2:
        afficher_etape_checklist()
    else:
        afficher_etape_rapport()

def afficher_etape_informations():
    st.title("📋 Informations sur le Fournisseur")
    st.markdown("---")
    
    col1, col2 = st.columns(2)
    
    with col1:
        st.subheader("Identification")
        nom_fournisseur = st.text_input("Nom du fournisseur*", 
                                         value=st.session_state.fournisseur_info.get("Nom du fournisseur", ""))
        adresse = st.text_area("Adresse de l'activité de production/fabrication*",
                               value=st.session_state.fournisseur_info.get("Adresse", ""))
        
        interlocuteurs = st.text_area("Interlocuteurs rencontrés + fonction/coordonnées*",
                                      value=st.session_state.fournisseur_info.get("Interlocuteurs", ""))
        
        effectif = st.text_input("Effectif total",
                                 value=st.session_state.fournisseur_info.get("Effectif", ""))
        
        qualite = st.text_input("Nombre de personnes au service qualité",
                                value=st.session_state.fournisseur_info.get("Service qualité", ""))
    
    with col2:
        st.subheader("Contact et Certifications")
        contact_crise = st.text_input("Contact en cas de crise sanitaire (retrait/rappel)*",
                                      value=st.session_state.fournisseur_info.get("Contact crise", ""))
        
        gamme_produits = st.text_area("Description de l'entreprise / gamme de produits livrés chez BIOCOOP*",
                                      value=st.session_state.fournisseur_info.get("Gamme produits", ""))
        
        annee_partenariat = st.text_input("Année de démarrage du partenariat avec Biocoop",
                                          value=st.session_state.fournisseur_info.get("Année partenariat", ""))
        
        certifications = st.text_area("Certifications qualité et produits (dont BIO et BIO COHÉRENCE) + dates de validité*",
                                      value=st.session_state.fournisseur_info.get("Certifications", ""))
        
        site_mixte = st.radio("Type de site", ["100% BIO", "Site mixte (BIO + conventionnel)"],
                             index=0 if st.session_state.fournisseur_info.get("Type site", "100% BIO") == "100% BIO" else 1)
    
    st.subheader("Informations complémentaires")
    col3, col4 = st.columns(2)
    
    with col3:
        date_visite = st.date_input("Date de l'audit*",
                                    value=datetime.now())
        auditeur = st.text_input("Nom de l'auditeur*",
                                 value=st.session_state.fournisseur_info.get("Auditeur", ""))
    
    with col4:
        magasin_referent = st.text_input("Magasin référent BIOCOOP",
                                         value=st.session_state.fournisseur_info.get("Magasin référent", ""))
        
        date_derniere_visite = st.text_input("Date de la dernière fiche de visite BIOCOOP",
                                            value=st.session_state.fournisseur_info.get("Dernière visite", ""))
    
    st.markdown("---")
    
    if st.button("➡️ Passer à la checklist d'audit", type="primary", use_container_width=True):
        # Sauvegarder les informations
        st.session_state.fournisseur_info = {
            "Nom du fournisseur": nom_fournisseur,
            "Adresse": adresse,
            "Interlocuteurs": interlocuteurs,
            "Effectif": effectif,
            "Service qualité": qualite,
            "Contact crise": contact_crise,
            "Gamme produits": gamme_produits,
            "Année partenariat": annee_partenariat,
            "Certifications": certifications,
            "Type site": site_mixte,
            "Date audit": date_visite.strftime("%d/%m/%Y"),
            "Auditeur": auditeur,
            "Magasin référent": magasin_referent,
            "Dernière visite": date_derniere_visite
        }
        st.session_state.current_step = 2
        st.rerun()

def afficher_etape_checklist():
    st.title("✅ Checklist d'Audit BIOCOOP")
    st.markdown("---")
    
    # Barre de progression
    total_items = sum(len(data["items"]) for data in CHECKLIST_AUDIT.values())
    items_completes = len([k for k in st.session_state.audit_data if st.session_state.audit_data[k].get("notation")])
    progress = items_completes / total_items if total_items > 0 else 0
    
    st.progress(progress, text=f"Progression : {items_completes}/{total_items} items complétés ({progress*100:.0f}%)")
    
    # Filtres
    col1, col2 = st.columns([3, 1])
    with col1:
        categorie_selectionnee = st.selectbox(
            "Sélectionner une catégorie",
            ["Toutes les catégories"] + list(CHECKLIST_AUDIT.keys())
        )
    
    with col2:
        filtre_notation = st.multiselect(
            "Filtrer par notation",
            ["A", "B", "C", "N/A"],
            default=[]
        )
    
    st.markdown("---")
    
    # Affichage des items
    categories_a_afficher = CHECKLIST_AUDIT.keys() if categorie_selectionnee == "Toutes les catégories" else [categorie_selectionnee]
    
    for categorie in categories_a_afficher:
        data = CHECKLIST_AUDIT[categorie]
        
        with st.expander(f"**{categorie}**", expanded=(categorie_selectionnee != "Toutes les catégories")):
            st.markdown(f"**Criticité**: {data['criticite']} | **Coefficient**: {data['coefficient']}x")
            st.markdown("---")
            
            for item in data["items"]:
                item_id = item["id"]
                
                # Initialiser les données si nécessaire
                if item_id not in st.session_state.audit_data:
                    st.session_state.audit_data[item_id] = {"notation": None, "commentaire": ""}
                
                # Appliquer le filtre
                if filtre_notation and st.session_state.audit_data[item_id]["notation"] not in filtre_notation:
                    continue
                
                # Conteneur pour chaque item
                with st.container():
                    st.markdown(f"**{item_id}** - {item['question']}")
                    st.caption(item['details'])
                    
                    col_note, col_comment = st.columns([1, 3])
                    
                    with col_note:
                        notation = st.selectbox(
                            "Notation",
                            options=list(NOTATION_OPTIONS.keys()),
                            format_func=lambda x: NOTATION_OPTIONS[x]["label"],
                            key=f"notation_{item_id}",
                            index=list(NOTATION_OPTIONS.keys()).index(st.session_state.audit_data[item_id]["notation"]) 
                                  if st.session_state.audit_data[item_id]["notation"] else 0,
                            label_visibility="collapsed"
                        )
                        st.session_state.audit_data[item_id]["notation"] = notation
                    
                    with col_comment:
                        commentaire = st.text_area(
                            "Commentaire / Constat",
                            value=st.session_state.audit_data[item_id]["commentaire"],
                            key=f"comment_{item_id}",
                            height=80,
                            placeholder="Détaillez vos observations, preuves, constats..."
                        )
                        st.session_state.audit_data[item_id]["commentaire"] = commentaire
                    
                    st.markdown("---")
    
    # Boutons de navigation
    col1, col2 = st.columns(2)
    with col1:
        if st.button("⬅️ Retour aux informations", use_container_width=True):
            st.session_state.current_step = 1
            st.rerun()
    
    with col2:
        if st.button("➡️ Générer le rapport", type="primary", use_container_width=True):
            st.session_state.current_step = 3
            st.rerun()

def afficher_etape_rapport():
    st.title("📊 Rapport d'Audit Final")
    st.markdown("---")
    
    # Calculer les scores
    score_global, details_categories = calculer_score_global(st.session_state.audit_data)
    niveau, couleur = get_niveau_conformite(score_global)
    
    # Affichage du score global
    col1, col2, col3 = st.columns(3)
    
    with col1:
        st.metric("Score Global", f"{score_global:.1f}%")
    
    with col2:
        st.metric("Niveau de Conformité", niveau)
    
    with col3:
        total_nc = sum(1 for v in st.session_state.audit_data.values() 
                      if v.get("notation") in ["B", "C"])
        st.metric("Non-conformités", total_nc)
    
    st.markdown("---")
    
    # Détails par catégorie
    st.subheader("📈 Scores par catégorie")
    
    df_scores = pd.DataFrame([
        {
            "Catégorie": cat.split(".")[1].strip(),
            "Score (%)": f"{info['score']:.1f}",
            "Criticité": info['criticite'],
            "Items évalués": info['items_evalues']
        }
        for cat, info in details_categories.items()
    ])
    
    st.dataframe(df_scores, use_container_width=True, hide_index=True)
    
    st.markdown("---")
    
    # Plan d'action (non-conformités)
    st.subheader("⚠️ Non-conformités identifiées")
    
    nc_list = []
    for categorie, data in CHECKLIST_AUDIT.items():
        for item in data["items"]:
            item_id = item["id"]
            if item_id in st.session_state.audit_data:
                notation = st.session_state.audit_data[item_id]["notation"]
                if notation in ["B", "C"]:
                    nc_list.append({
                        "ID": item_id,
                        "Catégorie": categorie.split(".")[1].strip(),
                        "Question": item["question"],
                        "Gravité": "Majeure" if notation == "C" else "Mineure",
                        "Commentaire": st.session_state.audit_data[item_id].get("commentaire", "")
                    })
    
    if nc_list:
        df_nc = pd.DataFrame(nc_list)
        st.dataframe(df_nc, use_container_width=True, hide_index=True)
    else:
        st.success("✅ Aucune non-conformité identifiée !")
    
    st.markdown("---")
    
    # Génération du rapport Excel
    st.subheader("📥 Télécharger le rapport complet")
    
    if not OPENPYXL_AVAILABLE:
        st.error("❌ La génération de rapport Excel n'est pas disponible. Le module openpyxl n'a pas pu être chargé.")
        st.info("💡 Vous pouvez copier les données ci-dessus ou contacter le support technique.")
    else:
        col1, col2 = st.columns(2)
        
        with col1:
            if st.button("🔄 Régénérer le rapport", use_container_width=True):
                st.rerun()
        
        with col2:
            buffer = generer_rapport_excel(st.session_state.fournisseur_info, st.session_state.audit_data)
            
            if buffer:
                nom_fichier = f"Audit_BIOCOOP_{st.session_state.fournisseur_info.get('Nom du fournisseur', 'Fournisseur').replace(' ', '_')}_{datetime.now().strftime('%Y%m%d')}.xlsx"
                
                st.download_button(
                    label="📥 Télécharger le rapport Excel",
                    data=buffer,
                    file_name=nom_fichier,
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    type="primary",
                    use_container_width=True
                )
    
    st.markdown("---")
    
    # Bouton retour
    if st.button("⬅️ Retour à la checklist", use_container_width=True):
        st.session_state.current_step = 2
        st.rerun()

if __name__ == "__main__":
    main()
